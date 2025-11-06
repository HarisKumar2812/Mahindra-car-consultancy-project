# import cv2
# from datetime import datetime
# from ultralytics import YOLO
# import easyocr
# import openpyxl
# from openpyxl.styles import Font
# import re
# import os

# # Load the model
# model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# # Initialize EasyOCR reader
# reader = easyocr.Reader(['en'])

# # Open the video file
# video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\newrequirements\vehicle02.mp4')

# # Get video properties
# frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
# frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
# fps = video.get(cv2.CAP_PROP_FPS)

# # Initialize variables for ROI selection
# selecting_roi = False
# roi_selected = False
# roi_start = (0, 0)
# roi_end = (0, 0)

# # Initialize dictionary to store license plate text with timing
# plate_text_dict = {}

# # Initialize dictionary to store previously detected plates
# previous_plates = {}

# # Excel file path
# excel_file = "plate_data.xlsx"

# # Check if the Excel file exists
# if os.path.exists(excel_file):
#     # Load the existing workbook and worksheet
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#     sno = worksheet.max_row  # Continue the serial number from the last entry
# else:
#     # Create a new workbook and worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

#     # Make headings bold
#     for cell in worksheet["1:1"]:
#         cell.font = Font(bold=True)

#     sno = 1

# # Define a function to validate license plate format
# def is_valid_license_plate(text):
#     # Adjust this pattern to match your specific license plate format
#     pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?([A-Z0-9-]\s?){2,6}$'
#     return bool(re.match(pattern, text))

# def draw_roi(event, x, y, flags, param):
#     global roi_start, roi_end, selecting_roi, roi_selected

#     if event == cv2.EVENT_LBUTTONDOWN:
#         roi_start = (x, y)
#         selecting_roi = True
#         roi_selected = False
#     elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
#         roi_end = (x, y)
#     elif event == cv2.EVENT_LBUTTONUP:
#         roi_end = (x, y)
#         selecting_roi = False
#         roi_selected = True

# # Create a window and set the mouse callback for ROI selection
# cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
# cv2.setMouseCallback("Select ROI", draw_roi)

# # Main loop to process video frames
# while True:
#     # Read a frame from the video
#     ret, frame = video.read()

#     # Break the loop if the video is over
#     if not ret:
#         break

#     # Check if ROI has been selected
#     if selecting_roi or roi_selected:
#         frame_copy = frame.copy()
#         cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
#         cv2.imshow("Select ROI", frame_copy)
#         roi_selected = False
#     else:
#         # Perform detection if ROI is selected
#         results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
#         for r in results:
#             if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
#                 for box in r.boxes.xyxy:
#                     x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
#                     # Ensure the coordinates are within the frame dimensions
#                     x1 = max(0, min(x1, frame_width))
#                     y1 = max(0, min(y1, frame_height))
#                     x2 = max(0, min(x2, frame_width))
#                     y2 = max(0, min(y2, frame_height))
#                     print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

#                     # Check if vehicle is within ROI
#                     if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
#                        (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
#                         # Crop the detected license plate region
#                         plate_region = frame[y1:y2, x1:x2]
#                         # Convert to grayscale
#                         gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
#                         # Perform OCR using EasyOCR
#                         plate_text = reader.readtext(gray_plate)
#                         if plate_text:
#                             current_time = datetime.now()
#                             current_plate = plate_text[0][1]
#                             print(f"Detected Text: {current_plate}")

#                             # Validate the detected license plate format
#                             if not is_valid_license_plate(current_plate):
#                                 print(f"Invalid Plate Format: {current_plate}")
#                                 continue

#                             # Check for duplicates
#                             if current_plate in previous_plates:
#                                 last_detection_time = previous_plates[current_plate]
#                                 if (current_time - last_detection_time).seconds < 10:
#                                     # Skip this detection as it's a duplicate
#                                     continue

#                             # Store the detection
#                             previous_plates[current_plate] = current_time
#                             plate_text_dict[current_time] = current_plate
#                             worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
#                             sno += 1
#                             print(f"Detected Plate: {current_plate} at {current_time}")

#                             # Draw the bounding box and the text on the frame
#                             cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#                             cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#     # Display the frame
#     cv2.imshow("Frame", frame)

#     # Check for 'q' key press to exit
#     key = cv2.waitKey(1)
#     if key == ord("q"):
#         break

#     # Check for 'c' key press to start detection when ROI is selected
#     if key == ord("c"):
#         selecting_roi = not selecting_roi

# # Release video capture and close all windows
# video.release()
# cv2.destroyAllWindows()

# # Save the Excel file
# workbook.save(excel_file)

# # Output the dictionary with timing and plate text
# print("Plate Text with Timing:")
# for time, plate_text in plate_text_dict.items():
#     print(f"{time}: {plate_text}")



# import cv2
# from datetime import datetime
# from ultralytics import YOLO
# import easyocr
# import openpyxl
# from openpyxl.styles import Font
# import re
# import os

# # Load the model
# model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# # Initialize EasyOCR reader
# reader = easyocr.Reader(['en'])

# # Open the video file
# video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\keerthana.mp4')

# # Get video properties
# frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
# frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
# fps = video.get(cv2.CAP_PROP_FPS)

# # Define the codec and create VideoWriter object
# output_video = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, (frame_width, frame_height))

# # Initialize variables for ROI selection
# selecting_roi = False
# roi_selected = False
# roi_start = (0, 0)
# roi_end = (0, 0)

# # Initialize dictionary to store license plate text with timing
# plate_text_dict = {}

# # Initialize dictionary to store previously detected plates
# previous_plates = {}

# # Excel file path
# excel_file = "plate_data.xlsx"

# # Check if the Excel file exists
# if os.path.exists(excel_file):
#     # Load the existing workbook and worksheet
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#     sno = worksheet.max_row  # Continue the serial number from the last entry
# else:
#     # Create a new workbook and worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

#     # Make headings bold
#     for cell in worksheet["1:1"]:
#         cell.font = Font(bold=True)

#     sno = 1

# # Define a function to validate license plate format
# def is_valid_license_plate(text):
#     # Adjust this pattern to match your specific license plate format
#     pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?([A-Z0-9-]\s?){2,6}$'
#     return bool(re.match(pattern, text))

# def draw_roi(event, x, y, flags, param):
#     global roi_start, roi_end, selecting_roi, roi_selected

#     if event == cv2.EVENT_LBUTTONDOWN:
#         roi_start = (x, y)
#         selecting_roi = True
#         roi_selected = False
#     elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
#         roi_end = (x, y)
#     elif event == cv2.EVENT_LBUTTONUP:
#         roi_end = (x, y)
#         selecting_roi = False
#         roi_selected = True

# # Create a window and set the mouse callback for ROI selection
# cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
# cv2.setMouseCallback("Select ROI", draw_roi)

# # Main loop to process video frames
# while True:
#     # Read a frame from the video
#     ret, frame = video.read()

#     # Break the loop if the video is over
#     if not ret:
#         break

#     # Check if ROI has been selected
#     if selecting_roi or roi_selected:
#         frame_copy = frame.copy()
#         cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
#         cv2.imshow("Select ROI", frame_copy)
#         roi_selected = False
#     else:
#         # Perform detection if ROI is selected
#         results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
#         for r in results:
#             if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
#                 for box in r.boxes.xyxy:
#                     x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
#                     # Ensure the coordinates are within the frame dimensions
#                     x1 = max(0, min(x1, frame_width))
#                     y1 = max(0, min(y1, frame_height))
#                     x2 = max(0, min(x2, frame_width))
#                     y2 = max(0, min(y2, frame_height))
#                     print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

#                     # Check if vehicle is within ROI
#                     if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
#                        (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
#                         # Crop the detected license plate region
#                         plate_region = frame[y1:y2, x1:x2]
#                         # Convert to grayscale
#                         gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
#                         # Perform OCR using EasyOCR
#                         plate_text = reader.readtext(gray_plate)
#                         if plate_text:
#                             current_time = datetime.now()
#                             current_plate = plate_text[0][1]
#                             print(f"Detected Text: {current_plate}")

#                             # Validate the detected license plate format
#                             if not is_valid_license_plate(current_plate):
#                                 print(f"Invalid Plate Format: {current_plate}")
#                                 continue

#                             # Check for duplicates
#                             if current_plate in previous_plates:
#                                 last_detection_time = previous_plates[current_plate]
#                                 if (current_time - last_detection_time).seconds < 10:
#                                     # Skip this detection as it's a duplicate
#                                     continue

#                             # Store the detection
#                             previous_plates[current_plate] = current_time
#                             plate_text_dict[current_time] = current_plate
#                             worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
#                             sno += 1
#                             print(f"Detected Plate: {current_plate} at {current_time}")

#                             # Draw the bounding box and the text on the frame
#                             cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#                             cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#                             # Show continuously detected text above the bounding box
#                             overlay_text = ' '.join([text[1] for text in plate_text])
#                             cv2.putText(frame, overlay_text, (x1, y1 - 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#     # Display the frame
#     cv2.imshow("Frame", frame)

#     # Write the frame to the output video
#     output_video.write(frame)

#     # Check for 'q' key press to exit
#     key = cv2.waitKey(1)
#     if key == ord("q"):
#         break

#     # Check for 'c' key press to start detection when ROI is selected
#     if key == ord("c"):
#         selecting_roi = not selecting_roi

# # Release video capture and close all windows
# video.release()
# output_video.release()
# cv2.destroyAllWindows()

# # Save the Excel file
# workbook.save(excel_file)

# # Output the dictionary with timing and plate text
# print("Plate Text with Timing:")
# for time, plate_text in plate_text_dict.items():
#     print(f"{time}: {plate_text}")


# import cv2
# from datetime import datetime
# from ultralytics import YOLO
# import easyocr
# import openpyxl
# from openpyxl.styles import Font
# import re
# import os

# # Load the model
# model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# # Initialize EasyOCR reader
# reader = easyocr.Reader(['en'])

# # Open the video file
# video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\keerthana.mp4')

# # Get video properties
# frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
# frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
# fps = video.get(cv2.CAP_PROP_FPS)

# # Define the codec and create VideoWriter object
# output_video = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, (frame_width, frame_height))

# # Initialize variables for ROI selection
# selecting_roi = False
# roi_selected = False
# roi_start = (0, 0)
# roi_end = (0, 0)

# # Initialize dictionary to store license plate text with timing
# plate_text_dict = {}

# # Initialize dictionary to store previously detected plates
# previous_plates = {}

# # Excel file path
# excel_file = "plate_data.xlsx"

# # Check if the Excel file exists
# if os.path.exists(excel_file):
#     # Load the existing workbook and worksheet
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#     sno = worksheet.max_row  # Continue the serial number from the last entry
# else:
#     # Create a new workbook and worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

#     # Make headings bold
#     for cell in worksheet["1:1"]:
#         cell.font = Font(bold=True)

#     sno = 1

# # Define a function to validate license plate format
# def is_valid_license_plate(text):
#     # Adjust this pattern to match your specific license plate format
#     pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?([A-Z0-9-]\s?){2,6}$'
#     return bool(re.match(pattern, text))

# # Define a function to clean OCR results
# def clean_license_plate_text(text):
#     # Convert 'O' to '0' when it's supposed to be a digit and vice versa
#     cleaned_text = []
#     for i, char in enumerate(text):
#         if char == 'O' and (i == 2 or i == 3):  # Position for district code digits
#             cleaned_text.append('0')
#         elif char == '0' and (i > 3 and i < len(text) - 4):  # Position for letters in the second part of the plate
#             cleaned_text.append('O')
#         else:
#             cleaned_text.append(char)
#     return ''.join(cleaned_text)

# def draw_roi(event, x, y, flags, param):
#     global roi_start, roi_end, selecting_roi, roi_selected

#     if event == cv2.EVENT_LBUTTONDOWN:
#         roi_start = (x, y)
#         selecting_roi = True
#         roi_selected = False
#     elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
#         roi_end = (x, y)
#     elif event == cv2.EVENT_LBUTTONUP:
#         roi_end = (x, y)
#         selecting_roi = False
#         roi_selected = True

# # Create a window and set the mouse callback for ROI selection
# cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
# cv2.setMouseCallback("Select ROI", draw_roi)

# # Main loop to process video frames
# while True:
#     # Read a frame from the video
#     ret, frame = video.read()

#     # Break the loop if the video is over
#     if not ret:
#         break

#     # Check if ROI has been selected
#     if selecting_roi or roi_selected:
#         frame_copy = frame.copy()
#         cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
#         cv2.imshow("Select ROI", frame_copy)
#         roi_selected = False
#     else:
#         # Perform detection if ROI is selected
#         results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
#         for r in results:
#             if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
#                 for box in r.boxes.xyxy:
#                     x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
#                     # Ensure the coordinates are within the frame dimensions
#                     x1 = max(0, min(x1, frame_width))
#                     y1 = max(0, min(y1, frame_height))
#                     x2 = max(0, min(x2, frame_width))
#                     y2 = max(0, min(y2, frame_height))
#                     print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

#                     # Check if vehicle is within ROI
#                     if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
#                        (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
#                         # Crop the detected license plate region
#                         plate_region = frame[y1:y2, x1:x2]
#                         # Convert to grayscale
#                         gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
#                         # Perform OCR using EasyOCR
#                         plate_text = reader.readtext(gray_plate)
#                         if plate_text:
#                             current_time = datetime.now()
#                             detected_plate = plate_text[0][1]
#                             current_plate = clean_license_plate_text(detected_plate)
#                             print(f"Detected Text: {current_plate}")

#                             # Validate the detected license plate format
#                             if not is_valid_license_plate(current_plate):
#                                 print(f"Invalid Plate Format: {current_plate}")
#                                 continue

#                             # Check for duplicates
#                             if current_plate in previous_plates:
#                                 last_detection_time = previous_plates[current_plate]
#                                 if (current_time - last_detection_time).seconds < 10:
#                                     # Skip this detection as it's a duplicate
#                                     continue

#                             # Store the detection
#                             previous_plates[current_plate] = current_time
#                             plate_text_dict[current_time] = current_plate
#                             worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
#                             sno += 1
#                             print(f"Detected Plate: {current_plate} at {current_time}")

#                             # Draw the bounding box and the text on the frame
#                             cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#                             cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#                             # Show continuously detected text above the bounding box
#                             overlay_text = ' '.join([text[1] for text in plate_text])
#                             cv2.putText(frame, overlay_text, (x1, y1 - 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#     # Display the frame
#     cv2.imshow("Frame", frame)

#     # Write the frame to the output video
#     output_video.write(frame)

#     # Check for 'q' key press to exit
#     key = cv2.waitKey(1)
#     if key == ord("q"):
#         break

#     # Check for 'c' key press to start detection when ROI is selected
#     if key == ord("c"):
#         selecting_roi = not selecting_roi

# # Release video capture and close all windows
# video.release()
# output_video.release()
# cv2.destroyAllWindows()

# # Save the Excel file
# workbook.save(excel_file)

# # Output the dictionary with timing and plate text
# print("Plate Text with Timing:")
# for time, plate_text in plate_text_dict.items():
#     print(f"{time}: {plate_text}")


# import cv2
# from datetime import datetime
# from ultralytics import YOLO
# import easyocr
# import openpyxl
# from openpyxl.styles import Font
# import re
# import os

# # Load the model
# model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# # Initialize EasyOCR reader
# reader = easyocr.Reader(['en'])

# # Open the video file
# video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\keerthana.mp4')

# # Get video properties
# frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
# frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
# fps = video.get(cv2.CAP_PROP_FPS)

# # Define the codec and create VideoWriter object
# output_video = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, (frame_width, frame_height))

# # Initialize variables for ROI selection
# selecting_roi = False
# roi_selected = False
# roi_start = (0, 0)
# roi_end = (0, 0)

# # Initialize dictionary to store license plate text with timing
# plate_text_dict = {}

# # Initialize dictionary to store previously detected plates
# previous_plates = {}

# # Excel file path
# excel_file = "plate_data.xlsx"

# # Check if the Excel file exists
# if os.path.exists(excel_file):
#     # Load the existing workbook and worksheet
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#     sno = worksheet.max_row  # Continue the serial number from the last entry
# else:
#     # Create a new workbook and worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

#     # Make headings bold
#     for cell in worksheet["1:1"]:
#         cell.font = Font(bold=True)

#     sno = 1

# # Define a function to validate license plate format
# def is_valid_license_plate(text):
#     # Adjust this pattern to match your specific license plate format
#     pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?[A-Z]{1,2}\s?[0-9]{1,5}$'
#     return bool(re.match(pattern, text))

# # Define a function to clean OCR results
# def clean_license_plate_text(text):
#     # Remove full stops
#     text = text.replace('.', '')
#     cleaned_text = []
#     for i, char in enumerate(text):
#         if char == 'O' and (i == 2 or i == 3):  # Position for district code digits
#             cleaned_text.append('0')
#         elif char == '0' and (i > 3):  # Position for letters in the second part of the plate
#             cleaned_text.append('O')
#         else:
#             cleaned_text.append(char)

#     cleaned_text = ''.join(cleaned_text)

#     # Ensure that after the first four characters, the sequence starts with a letter or space followed by a letter
#     if len(cleaned_text) > 4 and not cleaned_text[4].isalpha():
#         for j in range(4, len(cleaned_text)):
#             if cleaned_text[j].isalpha():
#                 cleaned_text = cleaned_text[:4] + ' ' + cleaned_text[j:]
#                 break

#     return cleaned_text

# def draw_roi(event, x, y, flags, param):
#     global roi_start, roi_end, selecting_roi, roi_selected

#     if event == cv2.EVENT_LBUTTONDOWN:
#         roi_start = (x, y)
#         selecting_roi = True
#         roi_selected = False
#     elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
#         roi_end = (x, y)
#     elif event == cv2.EVENT_LBUTTONUP:
#         roi_end = (x, y)
#         selecting_roi = False
#         roi_selected = True

# # Create a window and set the mouse callback for ROI selection
# cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
# cv2.setMouseCallback("Select ROI", draw_roi)

# # Main loop to process video frames
# while True:
#     # Read a frame from the video
#     ret, frame = video.read()

#     # Break the loop if the video is over
#     if not ret:
#         break

#     # Check if ROI has been selected
#     if selecting_roi or roi_selected:
#         frame_copy = frame.copy()
#         cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
#         cv2.imshow("Select ROI", frame_copy)
#         roi_selected = False
#     else:
#         # Perform detection if ROI is selected
#         results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
#         for r in results:
#             if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
#                 for box in r.boxes.xyxy:
#                     x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
#                     # Ensure the coordinates are within the frame dimensions
#                     x1 = max(0, min(x1, frame_width))
#                     y1 = max(0, min(y1, frame_height))
#                     x2 = max(0, min(x2, frame_width))
#                     y2 = max(0, min(y2, frame_height))
#                     print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

#                     # Check if vehicle is within ROI
#                     if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
#                        (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
#                         # Crop the detected license plate region
#                         plate_region = frame[y1:y2, x1:x2]
#                         # Convert to grayscale
#                         gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
#                         # Perform OCR using EasyOCR
#                         plate_text = reader.readtext(gray_plate)
#                         if plate_text:
#                             current_time = datetime.now()
#                             detected_plate = plate_text[0][1]
#                             current_plate = clean_license_plate_text(detected_plate)
#                             print(f"Detected Text: {current_plate}")

#                             # Validate the detected license plate format
#                             if not is_valid_license_plate(current_plate):
#                                 print(f"Invalid Plate Format: {current_plate}")
#                                 continue

#                             # Check for duplicates
#                             if current_plate in previous_plates:
#                                 last_detection_time = previous_plates[current_plate]
#                                 if (current_time - last_detection_time).seconds < 10:
#                                     # Skip this detection as it's a duplicate
#                                     continue

#                             # Store the detection
#                             previous_plates[current_plate] = current_time
#                             plate_text_dict[current_time] = current_plate
#                             worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
#                             sno += 1
#                             print(f"Detected Plate: {current_plate} at {current_time}")

#                             # Draw the bounding box and the text on the frame
#                             cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#                             cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#                             # Show continuously detected text above the bounding box
#                             overlay_text = ' '.join([text[1] for text in plate_text])
#                             cv2.putText(frame, overlay_text, (x1, y1 - 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#     # Display the frame
#     cv2.imshow("Frame", frame)

#     # Write the frame to the output video
#     output_video.write(frame)

#     # Check for 'q' key press to exit
#     key = cv2.waitKey(1)
#     if key == ord("q"):
#         break

#     # Check for 'c' key press to start detection when ROI is selected
#     if key == ord("c"):
#         selecting_roi = not selecting_roi

# # Release video capture and close all windows
# video.release()
# output_video.release()
# cv2.destroyAllWindows()

# # Save the Excel file
# workbook.save(excel_file)

# # Output the dictionary with timing and plate text
# print("Plate Text with Timing:")
# for time, plate_text in plate_text_dict.items():
#     print(f"{time}: {plate_text}")


# import cv2
# from datetime import datetime
# from ultralytics import YOLO
# import easyocr
# import openpyxl
# from openpyxl.styles import Font
# import re
# import os 

# # Load the model
# model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# # Initialize EasyOCR reader
# reader = easyocr.Reader(['en'])

# # Open the video file
# video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\newrequirements\just.mp4')

# # Get video properties
# frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
# frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
# fps = video.get(cv2.CAP_PROP_FPS)

# # Define the codec and create VideoWriter object
# output_video = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, (frame_width, frame_height))

# # Initialize variables for ROI selection
# selecting_roi = False
# roi_selected = False
# roi_start = (0, 0)
# roi_end = (0, 0)

# # Initialize dictionary to store license plate text with timing
# plate_text_dict = {}

# # Initialize dictionary to store previously detected plates
# previous_plates = {}

# # Excel file path
# excel_file = "plate_data.xlsx"

# # Check if the Excel file exists
# if os.path.exists(excel_file):
#     # Load the existing workbook and worksheet
#     workbook = openpyxl.load_workbook(excel_file)
#     worksheet = workbook.active
#     sno = worksheet.max_row  # Continue the serial number from the last entry
# else:
#     # Create a new workbook and worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

#     # Make headings bold
#     for cell in worksheet["1:1"]:
#         cell.font = Font(bold=True)

#     sno = 1

# # Define a function to validate license plate format
# def is_valid_license_plate(text):
#     # Adjust this pattern to match your specific license plate format
#     pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?[A-Z]{1,2}\s?[0-9]{1,4}$'
#     return bool(re.match(pattern, text))

# # Define a function to clean OCR results
# def clean_license_plate_text(text):
#     # Remove full stops
#     text = text.replace('.', '')
#     cleaned_text = []
#     for i, char in enumerate(text):
#         # Convert 'O' to '0' in the positions where numbers are expected
#         if char == 'O' and i in [2, 3, 6, 7]:  # Positions for numbers in the plate
#             cleaned_text.append('0')
#         else:
#             cleaned_text.append(char)

#     cleaned_text = ''.join(cleaned_text)

#     # Ensure that after the first four characters, the sequence starts with a letter or space followed by a letter
#     if len(cleaned_text) > 4 and not cleaned_text[4].isalpha():
#         for j in range(4, len(cleaned_text)):
#             if cleaned_text[j].isalpha():
#                 cleaned_text = cleaned_text[:4] + ' ' + cleaned_text[j:]
#                 break

#     return cleaned_text

# def draw_roi(event, x, y, flags, param):
#     global roi_start, roi_end, selecting_roi, roi_selected

#     if event == cv2.EVENT_LBUTTONDOWN:
#         roi_start = (x, y)
#         selecting_roi = True
#         roi_selected = False
#     elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
#         roi_end = (x, y)
#     elif event == cv2.EVENT_LBUTTONUP:
#         roi_end = (x, y)
#         selecting_roi = False
#         roi_selected = True

# # Create a window and set the mouse callback for ROI selection
# cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
# cv2.setMouseCallback("Select ROI", draw_roi)

# # Main loop to process video frames
# while True:
#     # Read a frame from the video
#     ret, frame = video.read()

#     # Break the loop if the video is over
#     if not ret:
#         break

#     # Check if ROI has been selected
#     if selecting_roi or roi_selected:
#         frame_copy = frame.copy()
#         cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
#         cv2.imshow("Select ROI", frame_copy)
#         roi_selected = False
#     else:
#         # Perform detection if ROI is selected
#         results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
#         for r in results:
#             if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
#                 for box in r.boxes.xyxy:
#                     x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
#                     # Ensure the coordinates are within the frame dimensions
#                     x1 = max(0, min(x1, frame_width))
#                     y1 = max(0, min(y1, frame_height))
#                     x2 = max(0, min(x2, frame_width))
#                     y2 = max(0, min(y2, frame_height))
#                     print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

#                     # Check if vehicle is within ROI
#                     if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
#                        (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
#                         # Crop the detected license plate region
#                         plate_region = frame[y1:y2, x1:x2]
#                         # Convert to grayscale
#                         gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
#                         # Perform OCR using EasyOCR
#                         plate_text = reader.readtext(gray_plate)
#                         if plate_text:
#                             current_time = datetime.now()
#                             detected_plate = plate_text[0][1]
#                             current_plate = clean_license_plate_text(detected_plate)
#                             print(f"Detected Text: {current_plate}")

#                             # Validate the detected license plate format
#                             if not is_valid_license_plate(current_plate):
#                                 print(f"Invalid Plate Format: {current_plate}")
#                                 continue

#                             # Check for duplicates
#                             if current_plate in previous_plates:
#                                 last_detection_time = previous_plates[current_plate]
#                                 if (current_time - last_detection_time).seconds < 10:
#                                     # Skip this detection as it's a duplicate
#                                     continue

#                             # Store the detection
#                             previous_plates[current_plate] = current_time
#                             plate_text_dict[current_time] = current_plate
#                             worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
#                             sno += 1
#                             print(f"Detected Plate: {current_plate} at {current_time}")

#                             # Draw the bounding box and the text on the frame
#                             cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#                             cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#                             # Show continuously detected text above the bounding box
#                             overlay_text = ' '.join([text[1] for text in plate_text])
#                             cv2.putText(frame, overlay_text, (x1, y1 - 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

#     # Display the frame
#     cv2.imshow("Frame", frame)

#     # Write the frame to the output video
#     output_video.write(frame)

#     # Check for 'q' key press to exit
#     key = cv2.waitKey(1)
#     if key == ord("q"):
#         break

#     # Check for 'c' key press to start detection when ROI is selected
#     if key == ord("c"):
#         selecting_roi = not selecting_roi

# # Release video capture and close all windows
# video.release()
# output_video.release()
# cv2.destroyAllWindows()

# # Save the Excel file
# workbook.save(excel_file)

# # Output the dictionary with timing and plate text
# print("Plate Text with Timing:")
# for time, plate_text in plate_text_dict.items():
#     print(f"{time}: {plate_text}")

import cv2
from datetime import datetime
from ultralytics import YOLO
import easyocr
import openpyxl
from openpyxl.styles import Font
import re
import os

# Load the model
model = YOLO(r"D:\intern\codsoftintern\Bus-Time-Management-5\best.pt")

# Initialize EasyOCR reader
reader = easyocr.Reader(['en'])

# Open the video file
video = cv2.VideoCapture(r'D:\intern\codsoftintern\Bus-Time-Management-5\newrequirements\just.mp4')

# Get video properties
frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
fps = video.get(cv2.CAP_PROP_FPS)

# Define the codec and create VideoWriter object
output_video = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, (frame_width, frame_height))

# Initialize variables for ROI selection
selecting_roi = False
roi_selected = False
roi_start = (0, 0)
roi_end = (0, 0)

# Initialize dictionary to store license plate text with timing
plate_text_dict = {}

# Initialize dictionary to store previously detected plates
previous_plates = {}

# Excel file path
excel_file = "plate_data.xlsx"

# Check if the Excel file exists
if os.path.exists(excel_file):
    # Load the existing workbook and worksheet
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active
    sno = worksheet.max_row  # Continue the serial number from the last entry
else:
    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Sno.", "Number Plate", "Date", "Time"])

    # Make headings bold
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)

    sno = 1

# Define a function to validate license plate format
def is_valid_license_plate(text):
    # Adjust this pattern to match your specific license plate format
    pattern = r'^(AP|AR|AS|BR|CG|GA|GJ|HR|HP|JH|KA|KL|MP|MH|MN|ML|MZ|NL|OD|PB|RJ|SK|TN|TS|TR|UP|UK|WB|AN|CH|DD|DL|JK|LA|LD|PY)\s?[0-9]{2}\s?[A-Z]{1,2}\s?[0-9]{1,4}$'
    return bool(re.match(pattern, text))

# Define a function to clean OCR results
def clean_license_plate_text(text):
    # Remove full stops
    text = text.replace('.', '')
    cleaned_text = []
    for i, char in enumerate(text):
        # Convert 'O' to '0' in the positions where numbers are expected
        if char == 'O' and i in [2, 3, 6, 7]:  # Positions for numbers in the plate
            cleaned_text.append('0')
        else:
            cleaned_text.append(char)

    cleaned_text = ''.join(cleaned_text)

    # Ensure that after the first four characters, the sequence starts with a letter or space followed by a letter
    if len(cleaned_text) > 4 and not cleaned_text[4].isalpha():
        for j in range(4, len(cleaned_text)):
            if cleaned_text[j].isalpha():
                cleaned_text = cleaned_text[:4] + ' ' + cleaned_text[j:]
                break

    return cleaned_text

def draw_roi(event, x, y, flags, param):
    global roi_start, roi_end, selecting_roi, roi_selected

    if event == cv2.EVENT_LBUTTONDOWN:
        roi_start = (x, y)
        selecting_roi = True
        roi_selected = False
    elif event == cv2.EVENT_MOUSEMOVE and selecting_roi:
        roi_end = (x, y)
    elif event == cv2.EVENT_LBUTTONUP:
        roi_end = (x, y)
        selecting_roi = False
        roi_selected = True

# Create a window and set the mouse callback for ROI selection
cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
cv2.setMouseCallback("Select ROI", draw_roi)

# Main loop to process video frames
while True:
    # Read a frame from the video
    ret, frame = video.read()

    # Break the loop if the video is over
    if not ret:
        break

    # Check if ROI has been selected
    if selecting_roi or roi_selected:
        frame_copy = frame.copy()
        cv2.rectangle(frame_copy, roi_start, roi_end, (0, 255, 0), 2)
        cv2.imshow("Select ROI", frame_copy)
        roi_selected = False
    else:
        # Perform detection if ROI is selected
        results = model.predict(frame, conf=0.5, show_labels=False, classes=0)
        for r in results:
            if r.boxes.xyxy.numel() > 0:  # check if tensor is not empty
                for box in r.boxes.xyxy:
                    x1, y1, x2, y2 = map(int, box.numpy())  # convert tensor to numpy array and map to int
                    # Ensure the coordinates are within the frame dimensions
                    x1 = max(0, min(x1, frame_width))
                    y1 = max(0, min(y1, frame_height))
                    x2 = max(0, min(x2, frame_width))
                    y2 = max(0, min(y2, frame_height))
                    print(f"Bounding Box Coordinates: x1={x1}, y1={y1}, x2={x2}, y2={y2}")

                    # Check if vehicle is within ROI
                    if (roi_start[0] <= x1 <= roi_end[0] or roi_start[0] <= x2 <= roi_end[0]) and \
                       (roi_start[1] <= y1 <= roi_end[1] or roi_start[1] <= y2 <= roi_end[1]):
                        # Crop the detected license plate region
                        plate_region = frame[y1:y2, x1:x2]
                        # Convert to grayscale
                        gray_plate = cv2.cvtColor(plate_region, cv2.COLOR_BGR2GRAY)
                        # Perform OCR using EasyOCR
                        plate_text = reader.readtext(gray_plate)
                        if plate_text:
                            current_time = datetime.now()
                            detected_texts = [text[1] for text in plate_text]
                            combined_text = ' '.join(detected_texts)
                            current_plate = clean_license_plate_text(combined_text)
                            print(f"Detected Text: {current_plate}")

                            # Validate the detected license plate format
                            if not is_valid_license_plate(current_plate):
                                print(f"Invalid Plate Format: {current_plate}")
                                continue

                            # Check for duplicates
                            if current_plate in previous_plates:
                                last_detection_time = previous_plates[current_plate]
                                if (current_time - last_detection_time).seconds < 10:
                                    # Skip this detection as it's a duplicate
                                    continue

                            # Store the detection
                            previous_plates[current_plate] = current_time
                            plate_text_dict[current_time] = current_plate
                            worksheet.append([sno, current_plate, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")])
                            sno += 1
                            print(f"Detected Plate: {current_plate} at {current_time}")

                            # Draw the bounding box and the text on the frame
                            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.putText(frame, current_plate, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

                            # Show continuously detected text above the bounding box
                            overlay_text = combined_text
                            cv2.putText(frame, overlay_text, (x1, y1 - 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    # Display the frame
    cv2.imshow("Frame", frame)

    # Write the frame to the output video
    output_video.write(frame)

    # Check for 'q' key press to exit
    key = cv2.waitKey(1)
    if key == ord("q"):
        break

    # Check for 'c' key press to start detection when ROI is selected
    if key == ord("c"):
        selecting_roi = not selecting_roi

# Release video capture and close all windows
video.release()
output_video.release()
cv2.destroyAllWindows()

# Save the Excel file
workbook.save(excel_file)

# Output the dictionary with timing and plate text
print("Plate Text with Timing:")
for time, plate_text in plate_text_dict.items():
    print(f"{time}: {plate_text}")
